from pathlib import Path
from datetime import date
import pandas as pd
from openpyxl import load_workbook

DATA_DIR = Path(
    "/Users/kelseyaguirre/Library/Mobile Documents/"
    "com~apple~CloudDocs/Documents/"
    "BrainInflammationCollaborative/Unhide/Mara_Columbia/"
    "Shared with Mara"
)

input_file = DATA_DIR / "4_23_26_SurveyandBFQDatav3.3.xlsx"

today = f"{date.today().month}_{date.today().day}_{date.today().strftime('%y')}"
output_file = DATA_DIR / f"{today}_SurveyandBFQDatav3.4.xlsx"

sheet_name = "Illness History"

# Put your 6 cells here using the Excel row number and column name
cells_to_clear = [
    {"row_number": 175, "column_name": "illnesshistory_explain"},
    {"row_number": 247, "column_name": "illnesshistory_explain"},
    {"row_number": 345, "column_name": "illnesshistory_explain"},
    {"row_number": 434, "column_name": "illnesshistory_explain"},
    {"row_number": 536, "column_name": "illnesshistory_explain"},
    {"row_number": 729, "column_name": "illnesshistory_explain"},
]

wb = load_workbook(input_file)
ws = wb[sheet_name]

# Map column names to Excel column numbers
headers = {
    cell.value: cell.column
    for cell in ws[1]
    if cell.value is not None
}

for item in cells_to_clear:
    row_number = item["row_number"]
    column_name = item["column_name"]

    if column_name not in headers:
        print(f"Column not found: {column_name}")
        continue

    col_number = headers[column_name]
    old_value = ws.cell(row=row_number, column=col_number).value

    ws.cell(row=row_number, column=col_number).value = None

    print(f"Cleared row {row_number}, column {column_name}")
    print(f"Old value: {old_value}")

wb.save(output_file)

print("\nDone")
print(f"Saved cleaned file to: {output_file}")